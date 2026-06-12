"""
テスト用の学生データエクセルファイルを作成するスクリプト
2シート構成で、2シート目のA列に学生IDを記載
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os

def create_student_excel():
    # ワークブックを作成
    wb = Workbook()

    # デフォルトのシートを削除
    wb.remove(wb.active)

    # シート1：概要
    sheet1 = wb.create_sheet("概要", 0)
    sheet1['A1'] = "学生写真埋め込み用エクセル"
    sheet1['A1'].font = Font(size=16, bold=True)
    sheet1['A3'] = "このファイルは学生写真を埋め込むためのテンプレートです。"
    sheet1['A4'] = "2シート目に学生IDと写真が埋め込まれます。"

    # シート2：学生データ
    sheet2 = wb.create_sheet("学生データ", 1)

    # ヘッダー行の設定
    headers = ["学生ID", "写真", "氏名", "学年"]
    for col, header in enumerate(headers, start=1):
        cell = sheet2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # テスト用の学生データ（10名分）
    student_data = [
        ("20240001", "山田太郎", "1年"),
        ("20240002", "佐藤花子", "1年"),
        ("20240003", "鈴木一郎", "2年"),
        ("20240004", "田中美咲", "2年"),
        ("20240005", "高橋健太", "3年"),
        ("20240006", "伊藤優子", "3年"),
        ("20240007", "渡辺翔太", "4年"),
        ("20240008", "中村さくら", "4年"),
        ("20240009", "小林大輔", "1年"),
        ("20240010", "加藤愛美", "2年"),
    ]

    # データ行の追加
    for row_idx, (student_id, name, grade) in enumerate(student_data, start=2):
        sheet2.cell(row=row_idx, column=1, value=student_id)  # A列: 学生ID
        # B列は写真用に空けておく
        sheet2.cell(row=row_idx, column=3, value=name)        # C列: 氏名
        sheet2.cell(row=row_idx, column=4, value=grade)       # D列: 学年

        # 行の高さを設定（写真が入る分）
        sheet2.row_dimensions[row_idx].height = 75

    # 列幅の設定
    sheet2.column_dimensions['A'].width = 12  # 学生ID
    sheet2.column_dimensions['B'].width = 15  # 写真
    sheet2.column_dimensions['C'].width = 15  # 氏名
    sheet2.column_dimensions['D'].width = 10  # 学年

    # 全セルを中央揃えに
    for row in sheet2.iter_rows(min_row=2, max_row=len(student_data)+1, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ファイルを保存
    output_path = "test_student_data.xlsx"
    wb.save(output_path)
    print(f"✓ テスト用エクセルファイルを作成しました: {output_path}")
    print(f"  - シート数: 2")
    print(f"  - 学生数: {len(student_data)}名")
    print(f"  - 学生ID: {student_data[0][0]} 〜 {student_data[-1][0]}")

if __name__ == "__main__":
    create_student_excel()
