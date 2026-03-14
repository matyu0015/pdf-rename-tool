#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
テキスト置換用のテストエクセルファイル作成スクリプト
"""

import openpyxl
from openpyxl.styles import Font, Alignment

def create_replacement_test_excel():
    """プレースホルダーを含むテストエクセルを作成"""

    # ワークブックを作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "イベント案内"

    # ヘッダー
    ws['A1'] = '案内文'
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center')

    # テストデータ（プレースホルダーを含む案内文）
    test_data = [
        "【重要】面接のご案内\n\n{{氏名}}様\n\nこの度は、書類選考を通過されましたことをお知らせいたします。\n\n■面接日時\n{{日時}}\n\n■実施方法\nオンライン面接（Zoom）\n\n■ZoomURL\n{{ZOOM_URL}}\n\n■面接官\n{{面接官}}",

        "{{氏名}}様へ\n\n次回面接のご案内です。\n\n日時: {{日時}}\n場所: {{場所}}\n担当: {{面接官}}\n\nご不明点がございましたらご連絡ください。",

        "面接日程確定のお知らせ\n\n{{氏名}}様\n\n以下の日程で面接を実施いたします。\n{{日時}}\n\nZoom URL: {{ZOOM_URL}}\nミーティングID: {{ミーティングID}}\n\nよろしくお願いいたします。",

        "【{{会社名}}】2次面接のご案内\n\n{{氏名}}様\n\n2次面接の日程が決定いたしました。\n\n日時: {{日時}}\n面接官: {{面接官}}\nZoom: {{ZOOM_URL}}",

        "オンライン説明会のご案内\n\n{{氏名}}様\n\n説明会を{{日時}}に開催いたします。\n\n参加URL: {{ZOOM_URL}}\n\nご参加をお待ちしております。",
    ]

    # データを挿入
    for i, text in enumerate(test_data, start=2):
        ws[f'A{i}'] = text
        ws[f'A{i}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[i].height = 150

    # 列幅を設定
    ws.column_dimensions['A'].width = 80

    # ファイルを保存
    filename = "test_replacement.xlsx"
    wb.save(filename)
    print(f"✓ テスト用エクセルを作成しました: {filename}")
    print(f"  - {len(test_data)}行のサンプルデータを含んでいます")
    print(f"  - プレースホルダー: {{氏名}}, {{日時}}, {{ZOOM_URL}}, {{面接官}}, {{場所}}, {{会社名}}, {{ミーティングID}}")
    return filename

if __name__ == '__main__':
    create_replacement_test_excel()
